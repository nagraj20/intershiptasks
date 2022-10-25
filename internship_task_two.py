import openpyxl

def choice_c():
    choice = int(input("enter the choice \n 1:full fixtures \n 2: Indias fixtures \n 3: Group 1 fixtures \n 4: Group 2 fixtures \n "))
    if choice == 1:
        fetch_all()
    elif choice == 2:
        india_fixture()
    elif choice == 3:
        group1_fixture()
    elif choice == 4:
        group2_fixture()
    else:
        print("Invalid choice")


def fetch_all():
    row = frame_name.max_row
    column = frame_name.max_column
    for i in range(1,row+1):
        print("match number ", end="")
        for j in range(1,column+1):
            print(frame_name.cell(i,j).value," ",end="")
        print()

def india_fixture():
    column = frame_name.max_column
    ls = [17,24,31,37,43]
    for i in ls:
        print("match number ",end="")
        for j in range(1,column+1):
            print(frame_name.cell(i,j).value," ",end="")
        print()

def group1_fixture():
    column = frame_name.max_column
    ls=[14,15,16,20,21,22,26,27,28,32,33,34,38,39,40]
    for i in ls:
        print("macth number:",end="")
        for j in range(1,column+1):
            print(frame_name.cell(i,j).value," ",end="")
        print()

def group2_fixture():
    column = frame_name.max_column
    ls = [17,18,19,23,24,25,29,30,31,35,36,37,41,42,43]
    for i in ls:
        print("macth number:",end="")
        for j in range(1,column+1):
            print(frame_name.cell(i,j).value," ",end="")
        print()



file = openpyxl.load_workbook('world_cup_fixture.xlsx')
frame_name = file['Sheet1']
choice_c()


